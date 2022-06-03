from flask import Flask, redirect, send_file, url_for, render_template, request, g, session
import mysql.connector
from dict import dbToUITableNameReplace
from dict import dbToUIColumnNameReplace
from dict import digiToDbTableNameReplace
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from copy import copy
from bearer import getBearerToken
import requests
from datetime import datetime

app = Flask(__name__)
app.secret_key = 'mybiggestsecret'

class User:
    def __init__(self, id, username, password):
        self.id = id
        self.username = username
        self.password = password
    
    def __repr__(self):
        return f'User: {self.username}, Password: {self.password}'

users = []
users.append(User(id=1, username='lynxal_team', password='lynxal2020'))

# DB connection params

host = 'localhost'
user = 'root'
password = 'somegoodpassword'   
db = ''

font = Font(name='Calibri', size=11,)

alignment=Alignment(horizontal='center', vertical='center',)

number_format = 'General'

digi_headers = {
    "Accept": "*/*",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
    "Content-Type": "application/json",
    "User-Agent": "python-requests/2.4.3 CPython/3.4.0",
    "X-DIGIKEY-Client-Id": "mEszZA7lW3tiCtvB8UAS4SlC8eDoGWPv",
    "Authorization": f"Bearer {getBearerToken()}",
}

mous_json = {
    "SearchByKeywordRequest": {
        "keyword": ""
    }
}

def getStocks():
    stock = request.form['stock']
    stocks = []
    if stock == 'main':
        stocks.append('Main')
    elif stock == 'production':
        stocks.append('Production')
    elif stock == 'prototyping':
        stocks.append('Prototyping')
    return stocks

def convertStockName(stock):
    if stock == 'main':
        return 'Main'
    elif stock == 'production':
        return 'Production'
    elif stock == 'prototyping':
        return 'Prototyping'
    elif stock == 'all':
        return 'All'
    else:
        return -1

def dbConnect():
    try:
        conn = mysql.connector.connect(host = host, user = user, password = password)
        return conn
    except Exception as e:
        redirect(url_for('genMessage', message = str(e)))

def getTables(cursor):
    stock = request.form['stock']
    if stock == 'main':
        db = 'main_stock'        
    elif stock == 'production':
        db = 'production_stock'       
    elif stock == 'prototyping':
        db = 'prototyping_stock'
    cursor.execute(f'USE {db}')
    getTablesCommand = 'SHOW TABLES'
    cursor.execute(getTablesCommand)
    tables = cursor.fetchall()
    return tables

def searchInAllTables(mpn):
    tableNames = []
    columnNames = []
    componentArray = []
    def appendTables(tableName):
        if tableName in dbToUITableNameReplace:
            for key in dbToUITableNameReplace:
                if tableName == key:
                    tableNames.append(dbToUITableNameReplace[key])
        else:
            tableNames.append(tableName.capitalize())
    def appendColumns(colName):
        if colName == 'StandardPackQty':
            ctNames.append('Reel Quantity')
            return
        if colName in dbToUIColumnNameReplace:
            for key in dbToUIColumnNameReplace:
                if colName == key:
                    ctNames.append(dbToUIColumnNameReplace[key])
        else:
            ctNames.append(colName)
    conn = dbConnect()
    cursor = conn.cursor()
    tables = getTables(cursor)
    for table in tables:
        query = f'SELECT * FROM {table[0]} WHERE ManufacturerPartNumber LIKE \'%{mpn}%\''
        cursor.execute(query)
        components = cursor.fetchall()
        if components:
            appendTables(table[0])
            ctNames = []
            compt = []
            getColumnNames = f'DESCRIBE {table[0]}'
            cursor.execute(getColumnNames)
            colNames = cursor.fetchall()
            for colName in colNames:
                appendColumns(colName[0])
            columnNames.append(ctNames)
            for component in components:
                compt.append(list(component))
            componentArray.append(compt)
    return tableNames, columnNames, componentArray

def searchExactMatchInAllTables(mpn):
    tableNames = []
    columnNames = []
    componentArray = []
    def appendTables(tableName):
        if tableName in dbToUITableNameReplace:
            for key in dbToUITableNameReplace:
                if tableName == key:
                    tableNames.append(dbToUITableNameReplace[key])
        else:
            tableNames.append(tableName.capitalize())
    def appendColumns(colName):
        if colName == 'StandardPackQty':
            ctNames.append('Reel Quantity')
            return
        if colName in dbToUIColumnNameReplace:
            for key in dbToUIColumnNameReplace:
                if colName == key:
                    ctNames.append(dbToUIColumnNameReplace[key])
        else:
            ctNames.append(colName)
    conn = dbConnect()
    cursor = conn.cursor()
    tables = getTables(cursor)
    for table in tables:
        query = f'SELECT * FROM {table[0]} WHERE ManufacturerPartNumber = \'{mpn}\''
        cursor.execute(query)
        components = cursor.fetchall()
        if components:
            appendTables(table[0])
            ctNames = []
            compt = []
            getColumnNames = f'DESCRIBE {table[0]}'
            cursor.execute(getColumnNames)
            colNames = cursor.fetchall()
            for colName in colNames:
                appendColumns(colName[0])
            columnNames.append(ctNames)
            for component in components:
                compt.append(component)
            componentArray.append(compt)
    return tableNames, columnNames, componentArray

def getQuantity(mpn):
    conn = dbConnect()
    cursor = conn.cursor()
    tables = getTables(cursor)
    for table in tables:
        query = f'SELECT Quantity FROM {table[0]} WHERE ManufacturerPartNumber = \'{mpn}\''
        cursor.execute(query)
        quantity = cursor.fetchall()
        if quantity:
            if len(quantity) > 1:
                qty = 0
                for i in range(len(quantity)):
                    qty += quantity[i][0]
                return qty
            return quantity[0][0]

def getId(cursor, table, mpn):
    getId = f'SELECT ID FROM {table} WHERE ManufacturerPartNumber = \'{mpn}\''
    cursor.execute(getId)
    Ids = cursor.fetchall()
    return Ids[0][0]

def getQuantityById(cursor, table, Id):
    getQuantity = f'SELECT Quantity FROM {table} WHERE ID = {Id}'
    cursor.execute(getQuantity)
    stockQuantities = cursor.fetchall()
    return stockQuantities[0][0]

def addById(cursor, table, Id, stockQuantity, qty):
    update = f'UPDATE {table} SET Quantity = ({stockQuantity} + {qty}) WHERE ID = {Id}'
    cursor.execute(update)

def add(mpn, qty):
    conn = dbConnect()
    cursor = conn.cursor()
    tables = getTables(cursor)
    found = False
    for table in tables:
        findmpn = f'SELECT * FROM {table[0]} WHERE ManufacturerPartNumber = \'{mpn}\''
        cursor.execute(findmpn)
        components = cursor.fetchall()
        if components:
            found = True
            Id = getId(cursor, table[0], mpn)
            stockQuantity = getQuantityById(cursor, table[0], Id)
            addById(cursor, table[0], Id, stockQuantity, qty)
            conn.commit()
            newStockQuantity = getQuantityById(cursor, table[0], Id)
            if newStockQuantity == stockQuantity + qty: 
                return 'Stock updated successfully!'
            else:
                return 'Something went wrong while updating the database!'
    if not found:
        r_digi = requests.get(f'https://api.digikey.com/Search/v3/Products/{mpn}', headers = digi_headers).json()
        try:
            if r_digi['ManufacturerPartNumber'] == mpn:
                componentType = r_digi['Family']['Value']
                if componentType in digiToDbTableNameReplace:
                    for key in digiToDbTableNameReplace:
                        if key == componentType:
                            insertInto = f"INSERT INTO {digiToDbTableNameReplace[key]} (ManufacturerPartNumber, Quantity, StandardPackQty, LastUpdated) VALUES ('{mpn}', {qty}, {r_digi['StandardPackage']}, '{datetime.today().strftime('%d/%m/%Y')}')"
                            cursor.execute(insertInto)
                            conn.commit()
                            Id = getId(cursor, digiToDbTableNameReplace[key], mpn)
                            stockQuantity = getQuantityById(cursor, digiToDbTableNameReplace[key], Id)
                            if stockQuantity == qty:
                                return 'Stock updated successfully!'
                            else:
                                return 'Something went wrong while updating the database!'
                else:
                    return 'App cannot understand the component type found, please select it manually'
            else:
                return 'Couldn\'t find the component type, please select it manually'
        except:
            return 'Couldn\'t find the component type, please select it manually'

def withdrawById(cursor, table, Id, stockQuantity, qty):
    update = f'UPDATE {table} SET Quantity = ({stockQuantity} - {qty}) WHERE ID = {Id}'
    cursor.execute(update)

def withdraw(mpn, qty):
    conn = dbConnect()
    cursor = conn.cursor()
    tables = getTables(cursor)
    found = False
    for table in tables:
        findmpn = f'SELECT * FROM {table[0]} WHERE ManufacturerPartNumber = \'{mpn}\''
        cursor.execute(findmpn)
        components = cursor.fetchall()
        if components:
            found = True
            Id = getId(cursor, table[0], mpn)
            stockQuantity = getQuantityById(cursor, table[0], Id)
            if stockQuantity >= qty:
                withdrawById(cursor, table[0], Id, stockQuantity, qty)
                conn.commit()
                newStockQuantity = getQuantityById(cursor, table[0], Id)
                if newStockQuantity == stockQuantity - qty: 
                    return 'Stock updated successfully!'
                else:
                    return 'Something went wrong while updating the database!'
            else:
                return 'Not enough to withdraw!'
    if not found:
        return 'Couldn\'t find the component!'

def calcReelQty(columns, components):
    for i in range(len(columns)):
        for index, columnName in enumerate(columns[i]):
            if columnName == 'Reel Quantity':
                for component in components[i]:
                    if component[index] != None:
                        if not (component[index - 1] == 0 or component[index] == 0):
                            if component[index - 1] % component[index] == 0:
                                component[index] = component[index - 1] // component[index]
                            else:
                                if round(component[index - 1] / component[index], 2) * 10 % 10 == 0:
                                    component[index] = int(round(component[index - 1] / component[index], 2) * 10 / 10)
                                else:
                                    component[index] = round(component[index - 1] / component[index], 2)
                        else:
                            component[index] = '-'
    

def getComponentLengths(components):
    componentLengths = []
    for i in range(len(components)):
        componentLengths.append(len(components[i]))
    return componentLengths

def getNumberOfColumns(columnNames):
    numberOfColumns = []
    for i in range(len(columnNames)):
        numberOfColumns.append(len(columnNames[i]))
    return numberOfColumns

def getExcelWbSheetsFilename():
    excel = request.files['excel']
    filename = excel.filename
    excel.save(filename)
    wb = load_workbook(filename, data_only=True)
    sheets = []
    for i in range (len(wb.worksheets)):
        sheets.append(wb.worksheets[i])
    return wb, sheets, filename

def getExcelColumn(sheet, columnName):
    for col in range(sheet.max_column):
        if sheet[1][col].value == columnName or sheet[1][col].value == columnName.lower() or sheet[1][col].value == columnName.upper():
            return col
    return None

@app.before_request
def before_request():
    g.user = None
    if 'user_id' in session:
        for i in range(len(users)):
            if users[i].id == session.get('user_id'):
                g.user = users[i]

@app.route('/sign_in', methods = ['POST', 'GET'])
def signIn():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        for i in range(len(users)):
            if username == users[i].username and password == users[i].password:
                session.permanent = True
                session['user_id'] = users[i].id
                return redirect(url_for('chooseAction'))
        return render_template('signin.html', failed = True)
    return render_template('signin.html')

@app.route('/sign_out', methods = ['GET'])
def signOut():
    if session.get('user_id'): 
        session.pop('user_id')
    return redirect(url_for('signIn'))

@app.route('/', methods = ['GET'])
def main():
    if not g.user:
        return redirect(url_for('signIn'))
    return redirect(url_for('chooseAction'))

@app.route('/choose_action', methods = ['POST', 'GET'])
def chooseAction():
    if not g.user:
        return redirect(url_for('signIn'))
    if request.method == 'POST':
        action = request.form['action']
        # if action == 'search' or action == 'inventorization' or action == 'add' or action == 'withdraw':  
        #     return redirect(url_for('chooseStock', action = action))
        if action == 'search' or action == 'withdraw' or action == 'add' or action == 'updateBOM':
            return redirect(url_for('chooseStock', action = action))
        # elif action == 'move': 
        #     return redirect(url_for('chooseStocksToMove'))
        else:
            return redirect(url_for('underDev'))
    return render_template('home.html')

@app.route('/choose_stock', methods = ['POST', 'GET'])
def chooseStock():
    if not g.user:
        return redirect(url_for('signIn'))
    if request.method == 'POST':
        stock = request.form['stock']
        if stock == 'all':
            return redirect(url_for('getInfo', action = request.form['action'], stock = stock))
        if stock == 'main' or stock == 'production' or stock == 'prototyping' or stock == 'ready_for_sale':
            return redirect(url_for('getInfo', action = request.form['action'], stock = stock))
    if request.args['action']:
        return render_template('Stocks/mainProdProt.html')

@app.route('/oops_bug', methods = ['GET'])
def underDev():
    return render_template('/Responses/underDev.html')

@app.route('/choose_stocks_to_move', methods = ['POST', 'GET'])
def chooseStocksToMove():
    if not g.user:
        return redirect(url_for('signIn'))
    if request.method == 'POST':
        pass
    return render_template('Stocks/moveStocks.html')

@app.route('/info_query', methods = ['GET'])
def getInfo():
    if not g.user:
        return redirect(url_for('signIn'))
    action = request.args.get('action')
    stock = convertStockName(request.args.get('stock'))
    if action == 'search':
        return render_template('Queries/search.html')
    elif action == 'withdraw':
        return render_template('Queries/withdraw.html')
    elif action == 'add':
        return render_template('Queries/add.html')
    elif action == 'updateBOM':
        return render_template('Queries/updateBOM.html', action = action, stock = stock)

@app.route('/search_by_mpn', methods = ['POST'])
def searchByMpn():
    if not g.user:
        return redirect(url_for('signIn'))
    stock = request.form['stock']
    mpn = request.form['mpn'].upper()
    stocks = getStocks()   
        #endregion
    if stock == 'all':
        # getDbs = 'SHOW DATABASES'
        # cursor.execute(getDbs)
        # dbs = cursor.fetchall()
        # for db in dbs:
        #     if db[0] != 'information_schema' and db[0] != 'mysql' and db[0] != 'performance_schema':
        #         use = f'USE {db[0]}'
        #         cursor.execute(use)
        #         getTables = 'SHOW TABLES'
        #         cursor.execute(getTables)
        #         tables = cursor.fetchall()
        #         for table in tables:
        #             query = f'SELECT * FROM {table[0]} WHERE ManufacturerPartNumber like \'{mpn}\''
        #             cursor.execute(query)
        #             components = cursor.fetchall()
        #             if components:
        #                 if db[0] == 'main_stock':
        #                     stockNames.append('Main')
        #                 elif db[0] == 'production_stock':
        #                     stockNames.append('Production')
        #                 elif db[0] == 'prototyping_stock':
        #                     stockNames.append('Prototyping')
        #                 appendTables(table[0])
        #             for component in components:
        #                 print(f'found in {db[0]}')
        #                 print(f'found in {table[0]}')
        #                 getColumnNames = f'SHOW COLUMNS FROM {table[0]}'
        #                 cursor.execute(getColumnNames)
        #                 colNames = cursor.fetchall()
        #                 for colName in colNames:
        #                     appendColumns(colName[0])
        #                 for param in component:
        #                     if param == 'None':
        #                         params.append('')
        #                     else:
        #                         params.append(param)
        #             columnNames.append(ctNames)
        #         print(columnNames)
        pass
    else:
        tables, columns, components = searchInAllTables(mpn)
        calcReelQty(columns, components)
    
    return render_template('Responses/search.html', stock = stock, mpn = mpn, stocks = stocks, tables = tables, columns = columns, numberOfColumns = getNumberOfColumns(columns), components = components, componentLengths = getComponentLengths(components))

@app.route('/search_by_values', methods = ['POST'])
def searchByValues():
    return redirect(url_for('underDev'))

@app.route('/search_by_file', methods = ['POST'])
def searchByFile():
    stock = request.form['stock']
    stocks = getStocks()
    sheets = getExcelWbSheetsFilename()[1]
    if 'Total_BOM' in sheets[0].title:
        sheets = sheets[slice(1, len(sheets), 1)]
    mpns = []
    for sheet in sheets:
        if getExcelColumn(sheet, 'Comment') != None:
            for row in range(2, sheet.max_row + 1):
                if sheet[row][getExcelColumn(sheet, 'Comment')].value != None:
                    mpns.append(sheet[row][getExcelColumn(sheet, 'Comment')].value)
    tables = []
    columns = []
    components = []
    for i in range (len(mpns)):
        tableNames, columnNames, comps = searchInAllTables(mpns[i])
        for tableName in tableNames:
            tables.append(tableName)
        for columnName in columnNames:
            columns.append(columnName)
        for comp in comps:
            components.append(comp)

    indicesOfSame = []
    for i in range(len(tables)):
        if tables[i] in tables[slice(i)]:
            continue
        for j in range(i + 1, len(tables)):
            if tables[i] == tables[j]:
                components[i] = components[i] + components[j]
                indicesOfSame.append(j)
    
    removed = 0
    for index in indicesOfSame:
        del tables[index - removed]
        del columns[index - removed]
        del components[index - removed]
        removed += 1

    calcReelQty(columns, components)

    return render_template('Responses/search.html', stock = stock, stocks = stocks, tables = tables, columns = columns, numberOfColumns = getNumberOfColumns(columns), components = components, componentLengths = getComponentLengths(components))

@app.route('/add_to_stock', methods = ['POST'])
def addToStock():
    mpn = request.form['mpn']
    qty = int(request.form['quantity'])
    return redirect(url_for('genMessage', message = add(mpn, qty)))

@app.route('/withdraw_from_stock', methods = ['POST'])
def withdrawFromStock():
    mpn = request.form['mpn']
    qty = int(request.form['quantity'])
    return redirect(url_for('genMessage', message = withdraw(mpn, qty)))

@app.route('/withdraw_by_file', methods = ['POST'])
def withdrawByFile():
    PCBquantity = int(request.form['quantity'])
    sheet = getExcelWbSheetsFilename()[1][0]
    mpns = []
    qtys = []
    msgs = []
    if getExcelColumn(sheet, "Comment") != None:
        for row in range(2, sheet.max_row + 1):
            if sheet[row][getExcelColumn(sheet, 'Comment')].value != None:
                mpns.append(sheet[row][getExcelColumn(sheet, 'Comment')].value)
                qtys.append(sheet[row][getExcelColumn(sheet, 'Quantity')].value)
    for i in range(len(mpns)):
        msgs.append(withdraw(mpns[i], PCBquantity * qtys[i]))

    notfoundMpns = []
    messages = []
    for i in range(len(msgs)):
        if msgs[i] == 'Something went wrong while updating the database!':
            notfoundMpns.append(mpns[i])
            messages.append('couldn\'t update!')
        elif msgs[i] == 'Not enough to withdraw!':
            notfoundMpns.append(mpns[i])
            messages.append('not enough to withdraw!')
        elif msgs[i] == 'Couldn\'t find the component!':
            notfoundMpns.append(mpns[i])
            messages.append('couldn\'t find the component!')

    return render_template('Responses/genMessage.html', message = 'Stock updated successfully!',  messages = messages, notfoundMpns = notfoundMpns)
    
@app.route('/update_bom_file', methods = ['POST'])
def updateBOM():
    wb, sheets, filename = getExcelWbSheetsFilename()
    if 'Total_BOM' in sheets[0].title:
        sheets = sheets[slice(1, len(sheets), 1)]

    mpn = ''

    mpns = []
    Lynxqtys = []
    Digiqtys = []
    Mousqtys = []

    for sheet in sheets:
        if not getExcelColumn(sheet, 'Digikey Stock'):
            sheet.cell(1, sheet.max_column + 1).value = 'Lynxal Stock'
            sheet.cell(1, sheet.max_column + 1).value = 'Digikey Stock'
            sheet.cell(1, sheet.max_column + 1).value = 'Mouser Stock'
            sheet[1][sheet.max_column - 3]._style = copy(sheet[1][sheet.max_column - 4]._style)
            sheet[1][sheet.max_column - 2]._style = copy(sheet[1][sheet.max_column - 4]._style)
            sheet[1][sheet.max_column - 1]._style = copy(sheet[1][sheet.max_column - 4]._style)
            sheet.column_dimensions[sheet[1][sheet.max_column - 3].coordinate[0]].bestFit = True
            sheet.column_dimensions[sheet[1][sheet.max_column - 2].coordinate[0]].bestFit = True
            sheet.column_dimensions[sheet[1][sheet.max_column - 1].coordinate[0]].bestFit = True

        for row in range (2, sheet.max_row + 1):
            sheet[row][getExcelColumn(sheet, 'Lynxal Stock')].font = font
            sheet[row][getExcelColumn(sheet, 'Lynxal Stock')].alignment = alignment
            sheet[row][getExcelColumn(sheet, 'Lynxal Stock')].number_format = number_format

            sheet[row][getExcelColumn(sheet, 'Digikey Stock')].font = font
            sheet[row][getExcelColumn(sheet, 'Digikey Stock')].alignment = alignment
            sheet[row][getExcelColumn(sheet, 'Digikey Stock')].number_format = number_format

            sheet[row][getExcelColumn(sheet, 'Mouser Stock')].font = font
            sheet[row][getExcelColumn(sheet, 'Mouser Stock')].alignment = alignment
            sheet[row][getExcelColumn(sheet, 'Mouser Stock')].number_format = number_format

            if not sheet[row][getExcelColumn(sheet, 'Comment')].value:
                continue
            
            mpn = sheet[row][getExcelColumn(sheet, 'Comment')].value
            requiredQty = sheet[row][getExcelColumn(sheet, 'Required Quantity')].value

            if mpn in mpns:
                index = mpns.index(mpn)
                if Lynxqtys[index] >= 1.1 * requiredQty:
                    sheet[row][getExcelColumn(sheet, 'Lynxal Stock')].value = Lynxqtys[index]
                else:
                    sheet[row][getExcelColumn(sheet, 'Lynxal Stock')].value = Lynxqtys[index]
                    sheet[row][getExcelColumn(sheet, 'Digikey Stock')].value = Digiqtys[index]
                    sheet[row][getExcelColumn(sheet, 'Mouser Stock')].value = Mousqtys[index]
            else:
                mpns.append(mpn)
                Lynxqty = getQuantity(mpn)
                if not Lynxqty:
                    Lynxqty = 0
                Lynxqtys.append(Lynxqty)
                sheet[row][getExcelColumn(sheet, 'Lynxal Stock')].value = Lynxqty
                if Lynxqty < 1.1 * requiredQty:
                    r_digi = requests.get(f'https://api.digikey.com/Search/v3/Products/{mpn}', headers = digi_headers).json()
                    mous_json['SearchByKeywordRequest']['keyword'] = mpn
                    r_mous = requests.post('https://api.mouser.com/api/v1/search/keyword?apiKey=0e7aaee3-b68a-4638-938b-c810074dc0d7', json = mous_json).json()
                    try:
                        if r_digi['ManufacturerPartNumber'] == mpn:
                            Digiqtys.append(r_digi['QuantityAvailable'])
                            sheet[row][getExcelColumn(sheet, 'Digikey Stock')].value = r_digi['QuantityAvailable']
                        else:
                            Digiqtys.append(0)
                            sheet[row][getExcelColumn(sheet, 'Digikey Stock')].value = 0
                    except:
                        Digiqtys.append(0)
                        sheet[row][getExcelColumn(sheet, 'Digikey Stock')].value = 0
                    try:
                        found = False
                        for i in range(len(r_mous['SearchResults']['Parts'])):
                            if r_mous['SearchResults']['Parts'][i]['ManufacturerPartNumber'] != mpn:
                                continue
                            found = True
                            if r_mous['SearchResults']['Parts'][i]['Availability'] == 'None' or 'On Order' in r_mous['SearchResults']['Parts'][i]['Availability']:
                                Mousqtys.append(0)
                                sheet[row][getExcelColumn(sheet, 'Mouser Stock')].value = 0
                            else:
                                Mousqtys.append(r_mous['SearchResults']['Parts'][i]['Availability'].replace(' In Stock', ''))
                                sheet[row][getExcelColumn(sheet, 'Mouser Stock')].value = r_mous['SearchResults']['Parts'][i]['Availability'].replace(' In Stock', '')
                            break
                        if not found:
                            Mousqtys.append(0)
                            sheet[row][getExcelColumn(sheet, 'Mouser Stock')].value = 0 
                    except:
                        Mousqtys.append(0)
                        sheet[row][getExcelColumn(sheet, 'Mouser Stock')].value = 0
                else:
                    Digiqtys.append('')
                    Mousqtys.append('')

    path = '(Updated) ' + filename
    wb.save(path)
    
    return send_file(path)

@app.route('/gen_message', methods = ['GET'])
def genMessage():
    return render_template('Responses/genMessage.html', message = request.args.get('message'))
    
        
if __name__  == "__main__":
    app.run(host='0.0.0.0', port=80, debug=True)